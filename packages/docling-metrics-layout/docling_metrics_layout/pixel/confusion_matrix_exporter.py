import colorsys
import logging
from pathlib import Path
from typing import Any, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pandas import ExcelWriter

from docling_metrics_layout.layout_types import (
    MultiLabelMatrixEvaluation,
)
from docling_metrics_layout.pixel.multi_label_confusion_matrix import (
    MultiLabelConfusionMatrix,
)

_log = logging.getLogger(__name__)


def linear_norm(x, x_min, x_max, k=5.0):
    d = x_max - x_min
    x = (x - x_min) / (x_max - x_min) if d != 0 else 0
    return x


def exp_norm(x, x_min, x_max, k=5.0):
    d = x_max - x_min
    x = (x - x_min) / (x_max - x_min) if d != 0 else 0
    return 1 - np.exp(-k * x)


def power_norm(x, x_min, x_max, p=0.3):
    d = x_max - x_min
    x = (x - x_min) / (x_max - x_min) if d != 0 else 0
    return x**p


def discover_filename_prefix(
    root: Path, filename_extension: str, scan_depth: int = 10
) -> Optional[str]:
    r"""
    Discover the common prefix used in the filenames of the prediction visualisations
    Check up to scan_depth files in the dir

    Return
    -------
    empty string: There is no common prefix
    None: Cannot reach consensus in any common prefix
    """

    def common_prefix(a: str, b: str, stop_char="_"):
        r"""String common prefix with stop char"""
        prefix: list[str] = []
        for c1, c2 in zip(a, b):
            if c1 != c2:
                break
            prefix.append(c1)
            if c1 == stop_char:
                break
        return "".join(prefix)

    prev_image_filename = None
    prefix = ""
    for i, image_fn in enumerate(root.glob(f"*.{filename_extension}")):
        if i >= scan_depth:
            break
        image_filename = image_fn.name
        if prev_image_filename:
            # new_prefix will be the empty string if there is nothing in common
            new_prefix = common_prefix(image_filename, prev_image_filename)
            if prefix == "":
                prefix = new_prefix
            elif new_prefix != "" and prefix != new_prefix:
                return None
        prev_image_filename = image_filename
    return prefix


class ConfusionMatrixExporter:
    r""" """

    TITLE_FONT_SIZE = 16
    SUBTITLE_FONT_SIZE = 14
    DATASET_WORKSHEET_NAME = "Dataset"
    IMAGES_WORKSHEET_NAME = "Images"

    def __init__(
        self,
    ):
        self._background_color = "bbbbbb"
        self._black_color = "444444"
        self._power_normalization_exp = 0.2

        border_style = "thick"
        border_color = "fc5a8d"
        self._highlighted_border = Border(
            left=Side(border_style=border_style, color=border_color),
            right=Side(border_style=border_style, color=border_color),
            top=Side(border_style=border_style, color=border_color),
            bottom=Side(border_style=border_style, color=border_color),
        )

    def build_ds_report(
        self,
        title: str,
        num_images: int,
        num_pixels: int,
        headers: list[str],
        matrix_evaluation: MultiLabelMatrixEvaluation,
        collapsed_headers: list[str],
        image_collaped_aggs: dict[str, np.ndarray],
        excel_fn: Path,
        visualisations_root: Optional[Path] = None,
    ):
        r"""
        Generate excel report for the full dataset
        """
        with pd.ExcelWriter(excel_fn, engine="openpyxl") as writer:
            # Add the dataset header
            wb: Workbook = writer.book
            if not wb.worksheets:
                wb.create_sheet(ConfusionMatrixExporter.DATASET_WORKSHEET_NAME)
                wb.active = 0
            ds_ws: Worksheet = wb.active  # type: ignore
            ds_ws.cell(row=1, column=1).value = title
            ds_ws.cell(row=1, column=1).font = Font(
                bold=True, size=ConfusionMatrixExporter.TITLE_FONT_SIZE
            )
            ds_ws.cell(row=2, column=1).value = "#images"
            ds_ws.cell(row=2, column=2).value = num_images
            ds_ws.cell(row=3, column=1).value = "#pixels"
            ds_ws.cell(row=3, column=2).value = num_pixels
            ds_ws.cell(row=3, column=2).number_format = "#,##0"

            # Build the basic report
            self._build_base_report(
                writer,
                ConfusionMatrixExporter.DATASET_WORKSHEET_NAME,
                headers,
                matrix_evaluation,
                4,
            )

            # Add the collapsed image metrics in a separate worksheet
            self._aggregate_collapsed_image_metrics(
                writer,
                ConfusionMatrixExporter.IMAGES_WORKSHEET_NAME,
                collapsed_headers,
                image_collaped_aggs,
                visualisations_root=visualisations_root,
            )

            # Adjust column widths
            images_ws: Worksheet = wb[ConfusionMatrixExporter.IMAGES_WORKSHEET_NAME]
            self._adjust_column_widths(ds_ws)
            self._adjust_column_widths(images_ws)

        _log.info("Dataset report: %s", str(excel_fn))

    def build_image_report(
        self,
        headers: list[str],
        matrix_evaluation: MultiLabelMatrixEvaluation,
        excel_fn: Path,
    ):
        with pd.ExcelWriter(excel_fn, engine="openpyxl") as writer:
            self._build_base_report(
                writer,
                ConfusionMatrixExporter.DATASET_WORKSHEET_NAME,
                headers,
                matrix_evaluation,
            )

            # Adjust column widths
            ws: Worksheet = writer.book[ConfusionMatrixExporter.DATASET_WORKSHEET_NAME]
            self._adjust_column_widths(ws)

        _log.info("Image report: %s", str(excel_fn))

    def _aggregate_collapsed_image_metrics(
        self,
        writer: ExcelWriter,
        worksheet_name: str,
        headers: list[str],
        image_collapsed_aggs: dict[str, np.ndarray],
        origin_cell: tuple[int, int] = (0, 0),
        decimal_digits: int = 3,
        visualisations_root: Optional[Path] = None,
    ):
        r"""
        Aggregate all collapsed image metrics
        """
        startrow = origin_cell[0] + 1
        startcol = origin_cell[1]

        # Build the dataframe
        index = list(image_collapsed_aggs.keys())
        data = np.stack(list(image_collapsed_aggs.values()), axis=0)  # [num_images, 12]
        data = np.round(data, decimals=3)
        df = pd.DataFrame(data, index=index, columns=headers)

        df.to_excel(
            writer,
            sheet_name=worksheet_name,
            index=True,
            startrow=startrow,
            startcol=startcol,
        )  # row/col index starts from 0

        # Load workbook
        wb: Workbook = writer.book
        ws: Worksheet = wb[worksheet_name]

        # Set the prediction visualisations as hyperlinks in the image filenames
        if visualisations_root:
            viz_prefix = discover_filename_prefix(visualisations_root, "png")
            if viz_prefix:
                col = startcol + 1
                for i, image_filename in enumerate(image_collapsed_aggs.keys()):
                    row = i + startrow + 2
                    cell = ws.cell(row=row, column=col)
                    viz_fn = visualisations_root / f"{viz_prefix}{image_filename}"
                    if not viz_fn.is_file():
                        continue
                    cell.hyperlink = str(viz_fn)
                    cell.style = "Hyperlink"
            else:
                _log.error(
                    "Cannot find any visualisation prefix in: %s",
                    str(visualisations_root),
                )

        # Set the subtitle
        subtitle_cell = ws.cell(
            row=origin_cell[0] + 1, column=origin_cell[1] + 1
        )  # start from 1
        subtitle_cell.value = "Image collapsed classes metrics"
        subtitle_cell.font = Font(
            bold=True, size=ConfusionMatrixExporter.SUBTITLE_FONT_SIZE
        )

        # Get data min/max values
        vmin: np.floating[Any] = np.min(data)
        vmax: np.floating[Any] = np.max(data)

        # Apply background colors to data cells
        row_start = 2 + startrow  # start from 1
        col_start = 2 + startcol
        for i in range(data.shape[0]):
            row = i + row_start
            for j in range(data.shape[1]):
                value = data[i, j]
                col = j + col_start
                if value == 0:
                    # Treat zero values specially
                    color = self._black_color
                else:
                    color = self._value_to_color(vmin, vmax, value, "linear")
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

                # Format the numbers
                decimals_format = "." + "0" * decimal_digits
                ws.cell(row=row, column=col).number_format = f"#,##0{decimals_format}"
        return ws.max_row, ws.max_column

    def _build_base_report(
        self,
        writer: ExcelWriter,
        worksheet_name: str,
        headers: list[str],
        matrix_evaluation: MultiLabelMatrixEvaluation,
        startrow: int = 0,
        hide_zero_rows: bool = True,
        hide_zero_cols: bool = True,
    ):
        r"""
        Generate excel report for a single image
        """
        detailed_spacing = 4  # spacing between the detailed matrices
        collapsed_spacing = (
            2  # spacing between a detailed and the next collapsed matrix
        )

        collapsed_headers = [
            headers[0],
            MultiLabelConfusionMatrix.ALL_COLLAPSED_CLASSES_NAME,
        ]

        # Add the confusion matrix
        max_row, max_col = self._export_matrix_to_excel(
            writer,
            worksheet_name,
            "Confusion Matrix",
            matrix_evaluation.detailed.confusion_matrix,
            headers,
            decimal_digits=3,
            origin_cell=(startrow, 0),
            normalization_func="linear",
            hide_zero_rows=hide_zero_rows,
            hide_zero_cols=hide_zero_cols,
        )

        # Add the precision matrix with detailed classes
        detailed_precision_row = max_row + detailed_spacing
        collapsed_precision_row = max_row + collapsed_spacing
        max_row, max_col = self._export_matrix_to_excel(
            writer,
            worksheet_name,
            "Precision Matrix",
            matrix_evaluation.detailed.precision_matrix,
            headers,
            decimal_digits=3,
            origin_cell=(detailed_precision_row, 0),
            normalization_func="linear",
            hide_zero_rows=hide_zero_rows,
            hide_zero_cols=hide_zero_cols,
        )
        detailed_recall_row = max_row + detailed_spacing
        collapsed_recall_row = max_row + collapsed_spacing
        collapsed_col = max_col + 1

        # Add the precision matrix with collapsed classes
        self._export_matrix_to_excel(
            writer,
            worksheet_name,
            "Collapsed Precision Matrix",
            matrix_evaluation.collapsed.precision_matrix,
            collapsed_headers,
            decimal_digits=3,
            origin_cell=(collapsed_precision_row, collapsed_col),
            normalization_func="linear",
            hide_zero_rows=hide_zero_rows,
            hide_zero_cols=hide_zero_cols,
        )

        # Add the recall matrix with detailed classes
        max_row, max_col = self._export_matrix_to_excel(
            writer,
            worksheet_name,
            "Recall matrix",
            matrix_evaluation.detailed.recall_matrix,
            headers,
            decimal_digits=3,
            origin_cell=(detailed_recall_row, 0),
            normalization_func="linear",
            hide_zero_rows=hide_zero_rows,
            hide_zero_cols=hide_zero_cols,
        )

        # Add the recall matrix with collapsed classes
        self._export_matrix_to_excel(
            writer,
            worksheet_name,
            "Collapsed Recall Matrix",
            matrix_evaluation.collapsed.recall_matrix,
            collapsed_headers,
            decimal_digits=3,
            origin_cell=(collapsed_recall_row, collapsed_col),
            normalization_func="linear",
            hide_zero_rows=hide_zero_rows,
            hide_zero_cols=hide_zero_cols,
        )

    def _export_matrix_to_excel(
        self,
        writer: ExcelWriter,
        worksheet_name: str,
        title: str,
        data: np.ndarray,
        headers: list[str],
        origin_cell: tuple[int, int] = (0, 0),
        special_first_cell: bool = False,
        decimal_digits: int = 0,
        normalization_func: str = "linear",  # One of 'linear', 'power', 'exp'
        hide_zero_rows: bool = False,
        hide_zero_cols: bool = False,
    ) -> tuple[int, int]:
        r"""
        Export the given data in excel and place it in the origin_cell

        Returns:
        --------
        max_row
        max_col
        """
        startrow = origin_cell[0]
        startcol = origin_cell[1]

        # Round values
        data = np.round(data, decimals=3)

        # Create DataFrame and write to Excel
        df = pd.DataFrame(data, index=headers, columns=headers)
        df.to_excel(
            writer,
            sheet_name=worksheet_name,
            index=True,
            startrow=startrow,
            startcol=startcol,
        )  # row/col index starts from 0

        # Load workbook
        wb: Workbook = writer.book
        ws: Worksheet = wb[worksheet_name]

        # Set the subtitle in the corner of the data
        subtitle_cell = ws.cell(
            row=origin_cell[0] + 1, column=origin_cell[1] + 1
        )  # start from 1
        subtitle_cell.value = title
        subtitle_cell.font = Font(
            bold=True, size=ConfusionMatrixExporter.SUBTITLE_FONT_SIZE
        )

        # Get the min/max values
        if special_first_cell:
            # Don't account for the first value
            confusion_mask = np.ones(data.shape, dtype=np.uint8)
            confusion_mask[0, 0] = 0
            vmin: np.floating[Any] = np.min(data, initial=0, where=confusion_mask != 0)
            vmax: np.floating[Any] = np.max(data, initial=0, where=confusion_mask != 0)
        else:
            vmin = np.min(data)
            vmax = np.max(data)

        # Apply background colors to data cells
        style_startrow = 2 + startrow  # start from 1
        style_startcol = 2 + startcol
        for i in range(len(headers)):
            row = i + style_startrow
            for j in range(len(headers)):
                col = j + style_startcol
                # Treat the background specially
                if i == 0 and j == 0 and special_first_cell:
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=self._background_color,
                        end_color=self._background_color,
                        fill_type="solid",
                    )
                else:
                    value = data[i, j]
                    if value == 0:
                        # Treat zero values specially
                        color = self._black_color
                    else:
                        color = self._value_to_color(
                            vmin, vmax, value, normalization_func
                        )
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

                # Highlight the diagonal
                if i == j:
                    ws.cell(row=row, column=col).border = self._highlighted_border

                # Format the numbers
                decimals_format = ""
                if decimal_digits > 0:
                    decimals_format = "." + "0" * decimal_digits
                ws.cell(row=row, column=col).number_format = f"#,##0{decimals_format}"

        # Hide rows/cols with all zeros
        if hide_zero_cols:
            colsums = np.sum(data, axis=0)
            zero_col_indices = np.nonzero(colsums == 0)[0]  # Zero column indices
            for zero_col_idx in zero_col_indices:
                col_idx = zero_col_idx + startcol + 2
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].hidden = True

        if hide_zero_rows:
            rowsums = np.sum(data, axis=1)
            zero_row_indices = np.nonzero(rowsums == 0)[0]  # Zero row indices
            for zero_row_idx in zero_row_indices:
                row_idx = zero_row_idx + startrow + 2
                ws.row_dimensions[row_idx].hidden = True

        return ws.max_row, ws.max_column

    def _value_to_color(self, vmin, vmax, v, normalization_func: str):
        """Map value to RGB color from blue→red using rainbow spectrum."""
        # Normalize to [0,1]
        if normalization_func == "power":
            normalized_value = power_norm(
                v, vmin, vmax, p=self._power_normalization_exp
            )
        elif normalization_func == "exp":
            normalized_value = exp_norm(v, vmin, vmax)
        else:
            normalized_value = linear_norm(v, vmin, vmax)

        # Use HSV rainbow mapping: hue 240° (blue) -> 0° (red)
        hue = (1 - normalized_value) * 240 / 360  # convert degrees to [0,1]
        r, g, b = colorsys.hsv_to_rgb(hue, 1, 1)

        # Convert to hex color for Excel
        hex_color = f"{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}"
        return hex_color

    def _adjust_column_widths(self, ws: Worksheet):
        r"""Adjust column widths for the final excel"""
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # type: ignore
            for cell in col:
                val = str(cell.value)
                max_length = max(max_length, len(val))
            ws.column_dimensions[col_letter].width = max_length + 2
